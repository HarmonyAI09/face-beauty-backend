from datetime import datetime
from typing import Optional
from fastapi.responses import FileResponse

from pydantic import EmailStr
from app.db.schemas.profile_schema import ProfileID, ProfileRead, ProfileReg
from app.db.session import get_db

from openpyxl import Workbook, load_workbook
import json

async def add_profile(info: ProfileReg) -> Optional[ProfileRead]:
    db = await get_db()
    profile_collection = db['profiles']
    MAX_COUNT = 5
    current_time = str((datetime.now()).date())

    existing_profile = await profile_collection.find_one({"mail": info.mail})

    if existing_profile:
        existing_ids = [ProfileID(id=item['id'], dateAdded=item['dateAdded'], name=item['name'], gender=item['gender'], racial=item['racial']) for item in existing_profile.get("profileIDs", [])]
        ids_set = {item.id for item in existing_ids}

        if len(existing_ids) < MAX_COUNT and info.profileID not in ids_set:
            await profile_collection.update_one(
                {"mail": info.mail},
                {
                    "$push": {"profileIDs": {"id": info.profileID, "dateAdded": current_time, "name": info.name, "gender": info.gender, "racial": info.racial}}
                }
            )
            updated_profile = await profile_collection.find_one({"mail": info.mail})
            updated_ids = [ProfileID(id=item['id'], dateAdded=item['dateAdded'], name=item['name'], gender=item['gender'], racial=item['racial']) for item in updated_profile.get("profileIDs", [])]
            return ProfileRead(mail=updated_profile["mail"], profileIDs=updated_ids)
    else:
        await profile_collection.insert_one({
            "mail": info.mail,
            "profileIDs": [{"id": info.profileID, "dateAdded": current_time, "name": info.name, "gender": info.gender, "racial": info.racial}]
        })
        new_profile = await profile_collection.find_one({"mail": info.mail})
        return ProfileRead(mail=new_profile["mail"], profileIDs=[ProfileID(**item) for item in new_profile.get("profileIDs", [])])

    return None

async def get_profile(mail: EmailStr) -> Optional[ProfileRead]:
    db = await get_db()
    profile_collection = db['profiles']
    profile_data = await profile_collection.find_one({"mail": mail})

    if not profile_data:
        return None

    profileIDs = [ProfileID(id=item['id'], dateAdded=item['dateAdded'], name=item['name'], gender=item['gender'], racial=item['racial']) for item in profile_data.get("profileIDs", [])]
    
    return ProfileRead(
        mail=profile_data["mail"],
        profileIDs=profileIDs
    )

async def get_report(id: str):
    db = await get_db()
    report_collection = db['store']
    report_data = await report_collection.find_one({"report_id": id})
    del report_data["_id"]

    return report_data

async def download_report(id: str):
    data = await get_report(id)
    #Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    #Write Info
    ws['A1'] = "Name"
    ws['A2'] = "Gender"
    ws['A3'] = "Ethnicity/Race"

    ws['B1'] = data["name"]
    ws['B2'] = data["gender"]
    ws['B3'] = data["race"]

    ws['A6'] = "Total"
    ws['A7'] = "Front"
    ws['A8'] = "Side"

    ws['B5'] = "Score"
    ws['C5'] = "Max Score"
    ws['D5'] = "Percentage"

    front = json.loads(data["front"]) if "front" in data else None
    side = json.loads(data["side"]) if "side" in data else None

    if front:
        ws['G5'] = "Front Profile"
        ws['G6'] = "No"
        ws['H6'] = "Feature"
        ws['I6'] = "Value"
        ws['J6'] = "Score"
        ws['K6'] = "Ideal Range"
        ws['L6'] = "Note"
        ws['M6'] = "Advice"

        for index in range(22):
            item = front['measurements'][index]
            ws['G' + str(7+index)] = index+1
            ws['H' + str(7+index)] = item['name']
            ws['H' + str(7+index)].hyperlink = "http://localhost:8000/get_image/"+data['report_id']+"/"+item['index']
            ws['I' + str(7+index)] = str(item['value'])
            ws['J' + str(7+index)] = item['score']
            ws['K' + str(7+index)] = str(item['ideal'])
            ws['L' + str(7+index)] = item['mean']
            ws['M' + str(7+index)] = item['advice']

    if side:
        ws['P5'] = "Side Profile"
        ws['P6'] = "No"
        ws['Q6'] = "Feature"
        ws['R6'] = "Value"
        ws['S6'] = "Score"
        ws['T6'] = "Ideal Range"
        ws['U6'] = "Note"
        ws['V6'] = "Advice"

        for index in range(23):
            item = side['measurements'][index]
            ws['P' + str(7+index)] = index+1
            ws['Q' + str(7+index)] = item['name']
            ws['Q' + str(7+index)].hyperlink = "http://localhost:8000/get_image/"+data['report_id']+"/"+str(index+22)
            ws['R' + str(7+index)] = str(item['value'])
            ws['S' + str(7+index)] = item['score']
            ws['T' + str(7+index)] = str(item['ideal'])
            ws['U' + str(7+index)] = item['mean']
            ws['V' + str(7+index)] = item['advice']

    ws['C7'] = 304.5
    ws['C8'] = 195.5
    ws['C6'] = '=C7+C8'
    ws['B6'] = '=B7+B8'
    ws['B7'] = '=SUM(J7:J28)'
    ws['B8'] = '=SUM(S7:S29)'
    ws['D6'] = '=B6/C6*100'
    ws['D7'] = '=B7/C7*100'
    ws['D8'] = '=B8/C8*100'


    wb.save(data['name']+'_'+data['race']+'_'+data['gender']+'.xlsx')
    file_path = data['name']+'_'+data['race']+'_'+data['gender']+'.xlsx'
    return FileResponse(file_path)